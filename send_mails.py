import os
import openpyxl
import subprocess
import time


def create_draft_with_attachment(subject, recipient, content, attachment_path):
    applescript = f"""
    tell application "Mail"
        set newMessage to make new outgoing message with properties {{subject:"{subject}", content:"{content}"}}
        tell newMessage
            make new to recipient at end of to recipients with properties {{address:"{recipient}"}}
            set theAttachment to POSIX file "{attachment_path}"
            make new attachment with properties {{file name:theAttachment}} at after the last paragraph
        end tell
        activate
    end tell
    """
    subprocess.run(["osascript", "-e", applescript])


def create_and_send_email(subject, recipient, content, attachment_path):
    applescript = f"""
    tell application "Mail"
        set newMessage to make new outgoing message with properties {{subject:"{subject}", content:"{content}"}}
        tell newMessage
            make new to recipient at end of to recipients with properties {{address:"{recipient}"}}
            set theAttachment to POSIX file "{attachment_path}"
            make new attachment with properties {{file name:theAttachment}} at after the last paragraph
            send
        end tell
        activate
    end tell
    """
    subprocess.run(["osascript", "-e", applescript])


def process_excel_files_in_directory(directory_path):
    for filename in os.listdir(directory_path):
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            file_path = os.path.join(directory_path, filename)

            # Open the Excel file and get the email address
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            email_address = sheet.cell(row=1, column=2).value
            name = sheet.cell(row=2, column=2).value
            grade = sheet.cell(row=35, column=2).value

            # Round the grade to the nearest 0.25
            rounded_grade = round(grade * 4) / 4

            # Define a dictionary for entertaining grade descriptions in German
            grade_descriptions = {
                6.00: "Sehr Gut - Fantastische Arbeit! 😃🌟👏",
                5.75: "Hervorragend - Unglaublich gut gemacht! Ihre Leistung ist beeindruckend! 👍🤩",
                5.50: "Hervorragend - Bemerkenswert! Sie zeigen grosses Potenzial! 👌😀",
                5.25: "Gut bis sehr gut - Wirklich toll gemacht! 🚀😎",
                5.00: "Gut - Tolle Arbeit! 👏😊",
                4.75: "Okay bis gut - Gute Arbeit! Bleiben Sie dran und Sie werden noch besser! 👍👨‍🏫",
                4.50: "Okay - Geschafft, aber mehr ist immer besser! Sie können noch mehr erreichen! 😅📚",
                4.25: "Genügend - Solide Leistung, aber da geht noch mehr 🤔✅",
                4.00: "Genügend - Sie haben es geschafft, aber da geht noch mehr 😐📖",
                3.75: "Ungenügend - Nahe dran, nicht ganz auf der Ziellinie 😕❌",
                3.50: "Ungenügend - Das geht besser! Lassen Sie uns mehr zusammenarbeiten, damit es besser läuft 😟❌",
                3.25: "Schwach - Sie haben das Potenzial, dich zu steigern! 😞❌",
                3.00: "Schwach - Da geht noch mehr! Gemeinsam werden wir besser! 😔❌",
            }

            # Check if the rounded grade is below 3
            if rounded_grade < 3:
                description = "Unglaublich schlecht - Das war nichts, aber beim nächsten Mal klappt's!"
            elif rounded_grade >= 3 and rounded_grade <= 6:
                description = grade_descriptions.get(
                    rounded_grade, "Ungültige Note - Hier stimmt etwas nicht."
                )
            else:
                description = "Ungültige Note - Hier stimmt etwas nicht."

            print(f"{name} ({email_address}), Note:  {grade}, Text: {description}")

            create_and_send_email(
                "Prüfung M290 - Resultat",
                email_address,
                f"{name} | Note {grade} | {description}",
                file_path,
            )
            time.sleep(2)


# Navigate to the "Results" directory and process the Excel files
results_directory = os.path.join(os.getcwd(), "Results")
process_excel_files_in_directory(results_directory)
